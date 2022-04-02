import argparse

from openpyxl import load_workbook
from PIL import Image


def hex_to_rgba(value: str) -> tuple:
    """Takes a hex value and returns a tuple of (r, g, b, a)

    Args:
        value (str): hex value

    Returns:
        tuple: tuple of (r, g, b, a)
    """
    value = value.lstrip('#')

    rgb = tuple(int(value[i:i + 2], 16) for i in range(0, len(value), 2))

    # Hex without alpha
    if len(rgb) == 3:
        return (*rgb, 255)

    # Hex with alpha
    return rgb


EXCELTHEME_COLOR = {
    0: '#ffffff',
    1: '#000000',
    4: '#4285f4',
    5: '#ea4335',
    6: '#fbbc04',
    7: '#34a853',
    8: '#ff6d01',
    9: '#46bdc6'
}


def xsl_to_png(xlsx_file: str, png_file: str = "output.png") -> None:
    """Converts an xlsx file to a png file

    Args:
        xsl_file (str): xsl file
        png_file (str): png file
    """
    wb = load_workbook(xlsx_file)
    ws = wb.active

    # Get the width and height of the sheet
    width = ws.max_column
    height = ws.max_row

    # Create a new image
    img = Image.new('RGBA', (int(width + 1), int(height + 1)))

    # Iterate over the cells
    for row in ws.rows:
        for cell in row:
            # Get the cell coordinates
            x = cell.column
            y = cell.row
            # Get the cell color
            color_obj = cell.fill.fgColor

            if color_obj.type == 'rgb':
                if len(color_obj.rgb) == 8:
                    hex_color = color_obj.rgb[2:] + color_obj.rgb[:2]
                rgba_color = hex_to_rgba(hex_color)
            elif color_obj.type == 'theme':
                if color_obj.theme in EXCELTHEME_COLOR:
                    hex_color = EXCELTHEME_COLOR[color_obj.theme]
                    rgba_color = hex_to_rgba(hex_color)
                else:  # Unknown theme
                    rgba_color = (0, 0, 0, 0)  # Transparent
            else:
                # Unknown color type
                rgba_color = (0, 0, 0, 0)  # Transparent
            # Set the pixel
            img.putpixel((int(x), int(y)), rgba_color)

    # Save the image
    img.save(png_file)


def main():
    """Main function"""
    # create parser object
    parser = argparse.ArgumentParser(
        description="A simple program to convert an xlsx file to a png file")

    # defining arguments for parser object
    parser.add_argument("-f", "--filepath", help="filepath of xlsx to convert", required=True)
    parser.add_argument("-o", "--output", help="png filename to save")

    # parse the arguments from standard input
    args = parser.parse_args()

    if args.output is not None:
        output_file = args.output
    else:
        output_file = args.filepath.strip(".xlsx") + ".png"

    xsl_to_png(args.filepath, output_file)


if __name__ == '__main__':
    main()
